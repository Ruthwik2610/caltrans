import os
import datetime

from dotenv import load_dotenv

# Load environment variables
load_dotenv()

import streamlit as st
from dotenv import load_dotenv
from openai import OpenAI
from PIL import Image

from src.chat_ui import text_based
from src.cucp_reevals import cucp_reevaluations
from src.foundation_model_chat import foundation_model_chat_ui
from src.highway_incident_summarizer import summarize_caltrans_incidents


def are_all_selected(options_list, selected_fields):
    return all(option in selected_fields for option in options_list)


# Set up the page layout
st.set_page_config(page_title="Caltrans", layout="wide")

# Apply blue sidebar styling
st.markdown(
    """
    <style>
    [data-testid="stSidebar"] {
        background-color: #3b82c8;
    }
    </style>
""",
    unsafe_allow_html=True,
)

# Load and apply custom CSS for general styling
with open("style/final.css") as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# Display company logo
st.image("image/caltrans.jpg", use_container_width=True)

# Header
st.markdown(
    "<p style='text-align: center; color: black; margin-top: -10px; font-size: 30px;'>"
    "<span style='font-weight: bold'>Agentic AI Learning Program – Foundational Learning</span></p>",
    unsafe_allow_html=True,
)

# Horizontal line
st.markdown(
    "<hr style='height: 2.5px; margin-top: 0px; width: 100%; background-color: gray; margin-left: auto; margin-right: auto;'>",
    unsafe_allow_html=True,
)

# Sidebar
with st.sidebar:
    st.markdown(
        "<p style='text-align: center; color: white; font-size:25px;'><span style='font-weight: bold; font-family: century-gothic';></span>Solutions Scope</p>",
        unsafe_allow_html=True,
    )
    vAR_AI_application = st.selectbox(
        "", ["Select Application", "Caltrans"], key="application"
    )
    vAR_LLM_model = st.selectbox(
        "",
        [
            "LLM Models",
            "gpt-3.5-turbo-16k-0613",
            "gpt-4-0314",
            "gpt-3.5-turbo-16k",
            "gpt-3.5-turbo-1106",
            "gpt-4-0613",
            "gpt-4-0314",
        ],
        key="text_llmmodel",
    )
    vAR_LLM_framework = st.selectbox(
        "", ["LLM Framework", "Langchain"], key="text_framework"
    )
    vAR_Gcp_cloud = st.selectbox(
        "",
        ["GCP Services Used", "VM Instance", "Computer Engine", "Cloud Storage"],
        key="text2",
    )
    st.markdown("#### ")
    href = """<form action="#">
        <input type="submit" value="Clear/Reset"/>
        </form>"""
    st.sidebar.markdown(href, unsafe_allow_html=True)
    st.markdown("# ")
    st.markdown(
        "<p style='text-align: center; color: White; font-size:20px;'>Build & Deployed on<span style='font-weight: bold'></span></p>",
        unsafe_allow_html=True,
    )
    s1, s2, s3 = st.columns((4, 4, 4))
    with s1:
        st.image("image/oie_png.png")
    with s2:
        st.image("image/aws_logo.png")
    with s3:
        st.image("image/AzureCloud_img.png")


# Layouts
col1, col2, col3, col4, col5 = st.columns((2, 5, 2, 5, 2))
col21, col22, col23, col24, col25 = st.columns((2, 5, 2, 5, 2))
col61, col62, col63, col64, col65 = st.columns((2, 5, 5, 5, 2))
col71, col72, col73 = st.columns([1, 7, 1])

# Only show usecase selection if application is selected
if vAR_AI_application == "Caltrans":
    with col2:
        st.subheader("Select Application")
        st.write("#")

    with col4:
        app_option = st.selectbox(
            "",
            (
                "Select the Usecase",
                "CUCP Re-Evaluations",
                "Foundation Model",
                "Guardrails",
                "Highway Incident Summarizer Bot",
                "Human in the feedback Loop",
                "Langchain",
                "LLM as a Judge",
                "LLM Training",
                "LLM Evaluation",
                "Prompt Engineering",
                "RAG-Document Intelligence",
                "Personal Narrative Insights",
            ),
            key="app_select",
        )
        st.write("## ")
else:
    # Show instruction to select application first
    with col2:
        st.subheader("Select Application")
        st.write("#")
    with col4:
        st.info("Please select the application from the sidebar to continue")
        app_option = "Select the Usecase"  # Default value

# Main Routing
if app_option != "Select the Usecase":
    # Handle external links - Auto-open in new tab
    if app_option == "Langchain":
        st.markdown(
            """
            <script>
                window.open('https://promptwithlangchain-398219119144.us-east1.run.app/', '_blank');
            </script>
        """,
            unsafe_allow_html=True,
        )

        st.info("🔗 Opening Langchain application in a new tab...")
        st.markdown(
            """
            <div style="margin-top: 15px;">
                <p style="font-size: 14px; color: #666; margin-bottom: 8px;">
                    If the new tab didn't open automatically:
                </p>
                <a href="https://promptwithlangchain-398219119144.us-east1.run.app/" target="_blank" rel="noopener noreferrer" style="text-decoration: none;">
                    <button style="
                        background: linear-gradient(135deg, #3b82c8 0%, #2563eb 100%);
                        color: white;
                        padding: 8px 16px;
                        border: none;
                        border-radius: 6px;
                        cursor: pointer;
                        font-size: 14px;
                        font-weight: 500;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                        transition: all 0.3s ease;
                    " onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 4px 8px rgba(0,0,0,0.15)';"
                    onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 2px 4px rgba(0,0,0,0.1)';">
                        Open Langchain App ↗
                    </button>
                </a>
            </div>
        """,
            unsafe_allow_html=True,
        )
        st.stop()

    elif app_option == "Prompt Engineering":
        st.markdown(
            """
            <script>
                window.open('https://genaiandpromptingtechnic-398219119144.us-east1.run.app/', '_blank');
            </script>
        """,
            unsafe_allow_html=True,
        )

        st.info("🔗 Opening Prompt Engineering application in a new tab...")
        st.markdown(
            """
            <div style="margin-top: 15px;">
                <p style="font-size: 14px; color: #666; margin-bottom: 8px;">
                    If the new tab didn't open automatically:
                </p>
                <a href="https://genaiandpromptingtechnic-398219119144.us-east1.run.app/" target="_blank" rel="noopener noreferrer" style="text-decoration: none;">
                    <button style="
                        background: linear-gradient(135deg, #3b82c8 0%, #2563eb 100%);
                        color: white;
                        padding: 8px 16px;
                        border: none;
                        border-radius: 6px;
                        cursor: pointer;
                        font-size: 14px;
                        font-weight: 500;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                        transition: all 0.3s ease;
                    " onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 4px 8px rgba(0,0,0,0.15)';"
                    onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 2px 4px rgba(0,0,0,0.1)';">
                        Open Prompt Engineering App ↗
                    </button>
                </a>
            </div>
        """,
            unsafe_allow_html=True,
        )
        st.stop()

    elif app_option == "Guardrails":
        text_based("Guardrails", "None")

    elif app_option in [
        "RAG-Document Intelligence",
        "Human in the feedback Loop",
    ]:
        with col22:
            st.subheader("Upload the Policy Document")
        with col24:
            knowledge_base = st.file_uploader(
                "Upload Knowledge Base",
                type=["pdf", "txt"],
                key="knowledge_base_upload",
            )
        text_based("RAG-Document Intelligence", knowledge_base)

    elif app_option == "LLM Training":
        with col22:
            st.subheader("Upload Training Data")
        with col24:
            training_file = st.file_uploader(
                "Upload Training Dataset (CSV, JSON, or JSONL)",
                type=["jsonl", "json", "csv"],
                key="training_file_upload",
            )
        text_based("LLM Training", training_file)

    elif app_option == "CUCP Re-Evaluations":
        with col22:
            st.subheader("Upload the Narrative(s) (PDF)")
        with col24:
            cucp_files = st.file_uploader(
                "Upload Narrative PDF(s)",
                type=["pdf"],
                key="cucp_upload",
                accept_multiple_files=True,
            )

        with col22:
            st.subheader("Upload PNW Data (Excel)")
        with col24:
            revenue_excel = st.file_uploader(
                "Upload PNW Data",
                type=["xlsx", "xls"],
                key="revenue_upload",
                help="Excel file with a 'Firms' sheet containing columns: Firm Name and Five Year Average",
            )

        # Parse the Excel file if provided
        firm_revenues = {}
        if revenue_excel:
            import pandas as pd
            import re
            try:
                # Read all sheets without assuming where the header is
                all_sheets = pd.read_excel(revenue_excel, sheet_name=None, header=None)
                
                target_df = None
                firm_col_name = None
                avg_col_name = None
                
                # Look specifically for the sheet containing "Firms" in its name (case-insensitive)
                sheet_names = list(all_sheets.keys())
                firms_sheet_name = next((name for name in sheet_names if "firms" in str(name).lower()), None)
                
                # If we found it, check that one. Otherwise, fallback to checking all of them.
                sheets_to_check = [firms_sheet_name] if firms_sheet_name else sheet_names
                
                for sheet_name in sheets_to_check:
                    if not sheet_name:
                        continue
                        
                    df = all_sheets[sheet_name]
                    
                    # Scan the first 15 rows to find the exact row with our headers
                    for i in range(min(15, len(df))):
                        # Convert to lowercase and normalize spaces (removes \n, \t, and extra spaces)
                        row_values = [re.sub(r'\s+', ' ', str(val)).strip().lower() for val in df.iloc[i].values]
                        
                        # Fuzzy matching: check if "firm" and "name" are in a cell, and "5" or "five" and "average" are in another
                        has_firm = any("firm" in v and "name" in v for v in row_values)
                        has_avg = any(("five" in v or "5" in v) and "average" in v for v in row_values)
                        
                        if has_firm and has_avg:
                            # Found the header! Set the columns and grab the data below it
                            df.columns = df.iloc[i]
                            target_df = df.iloc[i+1:].reset_index(drop=True)
                            
                            # Clean up column names internally to strip messy Excel spacing
                            cleaned_cols = [re.sub(r'\s+', ' ', str(c)).strip() for c in target_df.columns]
                            target_df.columns = cleaned_cols
                            
                            # Identify the exact column names the file uses to avoid KeyErrors
                            firm_col_name = next((c for c in cleaned_cols if "firm" in c.lower() and "name" in c.lower()), None)
                            avg_col_name = next((c for c in cleaned_cols if ("five" in c.lower() or "5" in c.lower()) and "average" in c.lower()), None)
                            break
                            
                    if target_df is not None and firm_col_name and avg_col_name:
                        break  # Stop looking once we've found our data
                
                # Now process the found data
                if target_df is not None and firm_col_name and avg_col_name:
                    target_df = target_df.dropna(subset=[firm_col_name])
                    
                    for _, row in target_df.iterrows():
                        firm_name = str(row[firm_col_name]).strip()
                        raw_revenue = row[avg_col_name]
                        
                        if firm_name.lower() == 'nan' or not firm_name:
                            continue
                            
                        # Handle formatting (e.g., "$1,234.50" -> float)
                        if isinstance(raw_revenue, str):
                            clean_rev = raw_revenue.replace("$", "").replace(",", "").strip()
                            try:
                                rev_val = float(clean_rev)
                            except ValueError:
                                rev_val = None
                        else:
                            try:
                                rev_val = float(raw_revenue)
                            except (ValueError, TypeError):
                                rev_val = None
                                
                        if rev_val is not None:
                            firm_revenues[firm_name] = rev_val
                else:
                    st.warning("Could not find columns resembling 'Firm Name' and 'Five Year Average'. Please ensure they exist in the 'Firms' sheet.")
            
            except Exception as e:
                st.error(f"Error parsing Excel file: {e}")

        # --- Memory Manager Sidebar ---
        with st.sidebar:
            
            # Push the feedback section lower in the sidebar
            st.markdown("###")
            st.markdown("###")
            st.markdown("---")
            
            from src.memory_manager import get_precedent_count, consolidate_memory_via_llm, overwrite_db
            
            total_precedents = get_precedent_count(1) + get_precedent_count(2) + get_precedent_count(3)
            pct = int((total_precedents / 135) * 100) if total_precedents > 0 else 0
            
            # Auto-consolidate if any level has hit the 45 limit
            any_level_full = any(get_precedent_count(lvl) >= 45 for lvl in [1, 2, 3])
            if any_level_full and not st.session_state.get('_auto_consolidated'):
                with st.spinner("Auto-consolidating corrections (a level reached its limit)..."):
                    clean_json_str = consolidate_memory_via_llm()
                    st.session_state['consolidated_rules_json'] = clean_json_str
                    st.session_state['show_consolidation_success'] = True
                    st.session_state['_auto_consolidated'] = True
            
            # --- Feedback + Merge Card ---
            # Inject a stable marker right before the container
            st.markdown('<div id="feedback-card-anchor"></div>', unsafe_allow_html=True)
            
            # CSS to style the adjacent container wrapper cleanly via sibling selector
            st.markdown("""
            <style>
            div[data-testid="stElementContainer"]:has(#feedback-card-anchor) + div[data-testid="stElementContainer"] div[data-testid="stVerticalBlockBorderWrapper"] {
                border: 2px solid rgba(255,255,255,0.35) !important;
                border-radius: 14px !important;
                background: rgba(255,255,255,0.06) !important;
                padding: 18px 16px !important;
            }
            </style>
            """, unsafe_allow_html=True)
            
            # Create a native Streamlit border container so it safely wraps the button
            feedback_card = st.container(border=True)
            
            with feedback_card:
                # Feedback Header
                st.markdown("""
                <div style="background: rgba(255,255,255,0.12); padding: 12px 16px; border-radius: 10px; text-align: center; margin-bottom: 14px;">
                    <p style="margin: 0; color: white; font-size: 1rem; font-weight: 600; letter-spacing: 0.04em; text-transform: uppercase;">Human in the Feedback Loop</p>
                </div>
                <p style="color: rgba(255,255,255,0.75); font-size: 0.82rem; margin: 0 0 14px 0; line-height: 1.5;">The AI learns from your corrections. Your edits are remembered and applied to all future evaluations.</p>
                """, unsafe_allow_html=True)
                
                # Metrics Area
                st.markdown(f"""
                <div style="background: rgba(255,255,255,0.08); border-radius: 10px; padding: 14px 16px; margin-bottom: 18px;">
                    <div style="display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 8px;">
                        <p style="margin: 0; color: rgba(255,255,255,0.7); font-size: 0.78rem; font-weight: 500; text-transform: uppercase;">Your Saved Corrections</p>
                        <p style="margin: 0; color: rgba(255,255,255,0.5); font-size: 0.72rem;">max 45 per step</p>
                    </div>
                    <p style="margin: 0 0 10px 0; color: white; font-size: 1.8rem; font-weight: 700;">{total_precedents}<span style="color: rgba(255,255,255,0.4); font-size: 1rem; font-weight: 400;"> / 135 total</span></p>
                    <div style="background: rgba(255,255,255,0.15); border-radius: 6px; height: 6px; overflow: hidden;">
                        <div style="background: linear-gradient(90deg, #60a5fa, #a78bfa); height: 100%; width: {pct}%; border-radius: 6px;"></div>
                    </div>
                    <p style="margin: 6px 0 0 0; color: rgba(255,255,255,0.5); font-size: 0.7rem; line-height: 1.4;">Each evaluation step can store up to 45 corrections. When full, corrections are auto-merged into a clean rulebook.</p>
                </div>
                """, unsafe_allow_html=True)
    
                # Separator
                st.markdown('<div style="height: 1px; background: rgba(255,255,255,0.12); margin: 0 0 14px 0;"></div>', unsafe_allow_html=True)
    
                # Merge Section
                st.markdown("""
                <div style="background: rgba(255,255,255,0.08); padding: 10px 14px; border-radius: 10px; text-align: center; margin-bottom: 8px;">
                    <p style="margin: 0; color: rgba(255,255,255,0.7); font-size: 0.88rem; font-weight: 600; text-transform: uppercase;">Merge Corrections</p>
                </div>
                <p style="color: rgba(255,255,255,0.55); font-size: 0.78rem; margin: 0 0 12px 0; line-height: 1.5;">Combine all corrections into a single clean rulebook file that you can download and re-upload later.</p>
                """, unsafe_allow_html=True)
                
                if total_precedents < 15:
                    st.markdown("""
                    <div style="
                        background: rgba(255,255,255,0.06);
                        padding: 10px 14px;
                        border-radius: 8px;
                        border: 1px solid rgba(255,255,255,0.1);
                        margin: 0 0 14px 0;
                    ">
                        <p style="margin: 0; color: rgba(255,255,255,0.65); font-size: 0.78rem; line-height: 1.5;">
                            Collect at least <strong style="color: rgba(255,255,255,0.9);">15 corrections</strong> before merging for best results.
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Button is now INSIDE the bordered card
                if st.button("Merge & Download Rulebook", use_container_width=True):
                    with st.spinner("AI is merging your corrections..."):
                        clean_json_str = consolidate_memory_via_llm()
                        st.session_state['consolidated_rules_json'] = clean_json_str
                        st.session_state['show_consolidation_success'] = True
                        
                if st.session_state.get('show_consolidation_success') and 'consolidated_rules_json' in st.session_state:
                    st.markdown("""
                    <div style="
                        background: rgba(74, 222, 128, 0.15);
                        padding: 10px 14px;
                        border-radius: 8px;
                        border: 1px solid rgba(74, 222, 128, 0.3);
                        margin: 10px 0;
                    ">
                        <p style="margin: 0; color: rgba(255,255,255,0.9); font-size: 0.82rem;">&#x2705; Merge complete. Download your rulebook below.</p>
                    </div>
                    """, unsafe_allow_html=True)
                    st.download_button(
                        label="Download Rulebook",
                        data=st.session_state['consolidated_rules_json'],
                        file_name=f"cucp_rules_{datetime.date.today().isoformat()}.json",
                        mime="application/json",
                        use_container_width=True
                    )

        if not cucp_files:
            st.markdown("""
            <div style="
                background: linear-gradient(135deg, #f0f7ff 0%, #e8f0fe 100%);
                border: 1px solid #bcd4f0;
                border-radius: 12px;
                padding: 24px 28px;
                margin: 20px 0;
            ">
                <h3 style="margin: 0 0 12px 0; color: #1a3d6e;">👋 Welcome! Here's how to get started:</h3>
                <ol style="margin: 0; padding-left: 20px; color: #333; line-height: 2;">
                    <li><strong>Upload the applicant's narrative</strong> (PDF) using the uploader above</li>
                    <li><strong>Optionally upload PNW data</strong> (Excel) for automatic cross-referencing</li>
                    <li>Click <strong>Start AI Evaluation</strong> — the AI will walk you through each step</li>
                </ol>
            </div>
            """, unsafe_allow_html=True)

        if cucp_files:
            cucp_file = cucp_files[0]  # Interactive mode processes one document at a time
            file_name = cucp_file.name
            base_name = file_name.rsplit(".", 1)[0]
            
            # --- Detect File Change ---
            if 'current_file_name' not in st.session_state:
                st.session_state.current_file_name = file_name
            elif st.session_state.current_file_name != file_name:
                # File changed! Wipe staged precedents and process state
                for key in ['eval_stage', 'pdf_text', 'l1_data', 'l2_data', 'l3_data', 'staged_precedents']:
                    if key in st.session_state:
                        del st.session_state[key]
                st.session_state.current_file_name = file_name
                st.rerun()
            
            # --- Initialize State Machine ---
            if 'eval_stage' not in st.session_state:
                st.session_state.eval_stage = 0
            if 'staged_precedents' not in st.session_state:
                st.session_state.staged_precedents = {
                    "level_1_precedents": [],
                    "level_2_precedents": [],
                    "level_3_precedents": []
                }
            if 'pdf_text' not in st.session_state:
                from src.cucp_reevals import extract_text_from_pdf
                st.session_state.pdf_text = extract_text_from_pdf(cucp_file)
            
            from src.cucp_reevals import run_level_1_extraction, run_level_2_classification, run_level_3_thresholds, generate_final_md_report
            from src.memory_manager import add_precedent
            
            # --- Progress Stepper ---
            step_labels = ["Start", "Step 1: Review Facts", "Step 2: Review Classifications", "Step 3: Review Thresholds", "Final Report"]
            current_step = st.session_state.eval_stage
            stepper_html = '<div style="display: flex; align-items: center; justify-content: center; margin: 10px 0 20px 0; gap: 0;">'
            for i, label in enumerate(step_labels):
                if i < current_step:
                    color = "#22c55e"; bg = "rgba(34,197,94,0.15)"; border_c = "#22c55e"; icon = "✓"
                elif i == current_step:
                    color = "#3b82f6"; bg = "rgba(59,130,246,0.15)"; border_c = "#3b82f6"; icon = str(i)
                else:
                    color = "#94a3b8"; bg = "rgba(148,163,184,0.08)"; border_c = "#cbd5e1"; icon = str(i)
                stepper_html += f'<div style="display:flex; flex-direction:column; align-items:center; min-width:80px;">'
                stepper_html += f'<div style="width:36px; height:36px; border-radius:50%; background:{bg}; border:2px solid {border_c}; display:flex; align-items:center; justify-content:center; font-size:0.85rem; font-weight:700; color:{color};">{icon}</div>'
                stepper_html += f'<p style="margin:4px 0 0 0; font-size:0.82rem; color:{color}; text-align:center; line-height:1.3;">{label}</p></div>'
                if i < len(step_labels) - 1:
                    line_color = "#22c55e" if i < current_step else "#cbd5e1"
                    stepper_html += f'<div style="flex:1; height:3px; background:{line_color}; margin:0 4px; margin-bottom:22px;"></div>'
            stepper_html += '</div>'
            st.markdown(stepper_html, unsafe_allow_html=True)
            
            # STATE 0: Upload & Process Level 1
            if st.session_state.eval_stage == 0:
                st.markdown("### Getting Started")
                with st.expander("📁 Have a saved rulebook from a previous session? Upload it here.", expanded=False):
                    uploaded_rules = st.file_uploader("Upload your previously saved corrections rulebook (optional)", type=['json'])
                    if uploaded_rules is not None:
                        try:
                            import json
                            rules_dict = json.load(uploaded_rules)
                            overwrite_db(rules_dict)
                            st.success("Successfully loaded your previous corrections!")
                        except Exception as e:
                            st.error(f"Failed to load rulebook: {e}")
                        
                st.markdown("---")
                if st.button("Start AI Evaluation ➔", type="primary"):
                    with st.spinner(f"Running Step 1: Reading and extracting facts from **{file_name}**..."):
                        l1_result = run_level_1_extraction(
                            st.session_state.pdf_text, 
                            firm_revenues, 
                            staged_precedents=st.session_state.staged_precedents.get("level_1_precedents", [])
                        )
                    if "error" in l1_result:
                        st.error(f"⚠️ Something went wrong during fact extraction: {l1_result['error']}")
                    else:
                        st.session_state.l1_data = l1_result
                        st.session_state.eval_stage = 1
                        st.rerun()
            
            # STATE 1: Level 1 Fact Review
            elif st.session_state.eval_stage == 1:
                st.markdown("### Step 1: Review Extracted Facts")
                st.info("The AI has read the narrative and pulled out the key facts. Please check if any facts are wrong or missing. If everything looks correct, click **Approve & Continue** at the bottom.")
                st.caption("*Tip: Double-click any cell in the table below to expand it and read the full text.*")
                
                # Display Data
                import pandas as pd
                l1_data = st.session_state.l1_data
                
                # Show Firm Info & Financials separately before the table
                st.markdown(f"**Found Applicant Firm:** `{l1_data.get('firm_name', 'None')}`")
                st.markdown(f"**Excel Cross-Reference Revenue/PNW:** `{l1_data.get('cross_reference_result', 'None')}`")
                st.markdown(f"**Narrative Declared PNW:** `{l1_data.get('narrative_pnw', 'NOT PROVIDED')}`")
                st.markdown("---")
                
                facts = l1_data.get('extracted_facts', [])
                df_facts = pd.DataFrame(facts)
                # Human-readable column headers
                col_rename = {
                    'id': 'Fact #', 'when': 'When', 'where': 'Where',
                    'who': 'Who', 'what': 'What', 'why': 'Why',
                    'magnitude': 'Magnitude/Threshold', 'demographic_flag': 'Demographics Checkbox',
                    'source_quote': 'Source Quote'
                }
                df_facts = df_facts.rename(columns={k: v for k, v in col_rename.items() if k in df_facts.columns})
                st.dataframe(df_facts, use_container_width=True)
                
                if st.button("⬅️ Go Back to Start"):
                    st.session_state.eval_stage = 0
                    st.rerun()
                
                # Correction Form (collapsed by default)
                l1_count = get_precedent_count(1) + len(st.session_state.staged_precedents.get("level_1_precedents", []))
                with st.expander("✏️ Something wrong? Click here to correct a fact", expanded=False):
                    with st.form("l1_correction_form", clear_on_submit=True):
                        st.markdown("#### Log a Fact Correction")
                        
                        if l1_count >= 45:
                            st.error(f"🚨 Correction limit reached ({l1_count}/45). Your corrections will be auto-merged. Check the sidebar to download the rulebook.")
                        elif l1_count >= 36:
                            st.warning(f"⚠️ Approaching correction limit ({l1_count}/45). Corrections will be auto-merged when the limit is reached.")
                            
                        c1, c2 = st.columns(2)
                        with c1:
                            target_fact = st.selectbox(
                                "Fact to Change (Context)", 
                                ["Firm Name", "Cross-Reference Result", "Narrative Declared PNW", "Demographics Checkbox", "Specific Incident Detail"],
                                help="Select the specific type of fact that the AI extracted incorrectly."
                            )
                        with c2:
                            correction_val = st.text_input(
                                "Corrected Value",
                                help="Enter the exact fact that the AI should have extracted."
                            )
                        reasoning_val = st.text_area(
                                "Reasoning for Change (What should the AI remember?)",
                                help="Explain why this is wrong. The AI will remember this and apply it to all future evaluations."
                            )
                        
                        # Disable submit if over threshold
                        is_disabled = True if l1_count >= 45 else False
                        if st.form_submit_button("Apply Correction & Re-Evaluate", disabled=is_disabled):
                            st.session_state.staged_precedents["level_1_precedents"].append({
                                "target": target_fact,
                                "correction": correction_val,
                                "human_reasoning": reasoning_val
                            })
                            # Auto re-run Level 1 with the new correction and stay on review
                            with st.spinner(f"Re-running Fact Extraction with your correction..."):
                                l1_result = run_level_1_extraction(
                                    st.session_state.pdf_text,
                                    firm_revenues,
                                    staged_precedents=st.session_state.staged_precedents.get("level_1_precedents", [])
                                )
                                st.session_state.l1_data = l1_result
                            st.rerun()
                
                st.caption("*Your corrections are saved only after you approve the final evaluation at the end.*")
                
                # Proceed Button
                excel_pnw = st.session_state.l1_data.get("cross_reference_result", "None")
                narrative_pnw = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                combined_financials = f"Excel Cross-Reference Revenue/PNW: {excel_pnw}\nNarrative Declared PNW: {narrative_pnw}"
                
                if st.button("Approve & Continue ➔", type="primary"):
                    with st.spinner("Running Step 2: Legal Classification..."):
                        l2_result = run_level_2_classification(
                            st.session_state.l1_data.get('extracted_facts', []),
                            combined_financials,
                            staged_precedents=st.session_state.staged_precedents.get("level_2_precedents", [])
                        )
                    if "error" in l2_result:
                        st.error(f"⚠️ Something went wrong during classification: {l2_result['error']}")
                    else:
                        st.session_state.l2_data = l2_result
                        st.session_state.eval_stage = 2
                        st.rerun()

            # STATE 2: Level 2 Classification Review
            elif st.session_state.eval_stage == 2:
                st.markdown("### Step 2: Review Legal Classifications")
                st.info("The AI has categorized each fact under 49 CFR §26.67 (e.g., Systemic Barrier, Economic Hardship). Review the categories below. If a fact is in the wrong category, click the correction section to fix it. Otherwise, click **Approve & Continue**.")
                st.caption("*Tip: Double-click any cell in the table below to expand it and read the full text.*")
                
                l2_data = st.session_state.l2_data
                classifications = l2_data.get('classifications', [])
                df_class = pd.DataFrame(classifications)
                # Human-readable column headers
                l2_col_rename = {
                    'fact_id': 'Fact #', 'classification': 'Legal Category',
                    'reasoning': 'AI Reasoning'
                }
                df_class = df_class.rename(columns={k: v for k, v in l2_col_rename.items() if k in df_class.columns})
                st.dataframe(df_class, use_container_width=True)
                
                if st.button("⬅️ Go Back to Step 1"):
                    st.session_state.eval_stage = 1
                    st.rerun()
                
                l2_count = get_precedent_count(2) + len(st.session_state.staged_precedents.get("level_2_precedents", []))
                with st.expander("✏️ Wrong category? Click here to reclassify a fact", expanded=False):
                    with st.form("l2_correction_form", clear_on_submit=True):
                        st.markdown("#### Override a Classification")
                        
                        if l2_count >= 45:
                            st.error(f"🚨 Correction limit reached ({l2_count}/45). Your corrections will be auto-merged. Check the sidebar to download the rulebook.")
                        elif l2_count >= 36:
                            st.warning(f"⚠️ Approaching correction limit ({l2_count}/45). Corrections will be auto-merged when the limit is reached.")
                            
                        c1, c2 = st.columns(2)
                        with c1:
                            target_class = st.selectbox(
                                "Scenario to Reclassify", 
                                [c.get('classification', c.get('Legal Category', '')) + ' (Change this)' for c in classifications] + ["General Scenario Misclassification"],
                                help="Select the specific AI classification you want to override."
                            )
                        with c2:
                            correction_class = st.selectbox(
                                "New Category", 
                                ["Systemic Barrier", "Ordinary Business Risk", "Economic Hardship", "Not a SED Barrier", "Other"],
                                help="Select the correct legal classification for this fact."
                            )
                        reasoning_class = st.text_area(
                                "Legal Rationale (Why is this the correct classification?)",
                                help="Provide your legal justification. The AI will learn this and apply it to similar cases in the future."
                            )
                        
                        is_disabled = True if l2_count >= 45 else False
                        if st.form_submit_button("Apply Override & Re-Evaluate", disabled=is_disabled):
                            st.session_state.staged_precedents["level_2_precedents"].append({
                                "target": target_class,
                                "correction": correction_class,
                                "human_reasoning": reasoning_class
                            })
                            # Re-run with FULL financial context preserved
                            excel_pnw_rerun = st.session_state.l1_data.get("cross_reference_result", "None")
                            narrative_pnw_rerun = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                            combined_financials_rerun = f"Excel Cross-Reference Revenue/PNW: {excel_pnw_rerun}\nNarrative Declared PNW: {narrative_pnw_rerun}"
                            with st.spinner("Re-evaluating Classifications..."):
                                st.session_state.l2_data = run_level_2_classification(
                                    st.session_state.l1_data.get('extracted_facts', []),
                                    combined_financials_rerun,
                                    staged_precedents=st.session_state.staged_precedents.get("level_2_precedents", [])
                                )
                            st.rerun()
                            
                st.caption("*Your corrections are saved only after you approve the final evaluation at the end.*")

                excel_pnw = st.session_state.l1_data.get("cross_reference_result", "None")
                narrative_pnw = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                combined_financials = f"Excel Cross-Reference Revenue/PNW: {excel_pnw}\nNarrative Declared PNW: {narrative_pnw}"

                if st.button("Approve & Continue ➔", type="primary"):
                    with st.spinner("Running Step 3: Evidentiary Thresholds..."):
                        l3_result = run_level_3_thresholds(
                            st.session_state.l2_data.get("classifications", []), 
                            st.session_state.l1_data.get("extracted_facts", []), 
                            combined_financials,
                            staged_precedents=st.session_state.staged_precedents.get("level_3_precedents", [])
                        )
                    if "error" in l3_result:
                        st.error(f"⚠️ Something went wrong during threshold evaluation: {l3_result['error']}")
                    else:
                        st.session_state.l3_data = l3_result
                        st.session_state.eval_stage = 3
                        st.rerun()

            # STATE 3: Level 3 Threshold Review
            elif st.session_state.eval_stage == 3:
                st.markdown("### Step 3: Review Pass/Fail Decisions")
                st.info("The AI has evaluated the 7 mandatory CUCP criteria and assigned Pass or Fail to each. Review the decisions below. If the AI is too strict or too lenient on any criterion, click the correction section. Otherwise, click **Approve & Finalize**.")
                st.caption("*Tip: Double-click any cell in the table below to expand it and read the full text.*")
                
                l3_data = st.session_state.l3_data
                criteria = l3_data.get('criteria', [])
                df_crit = pd.DataFrame(criteria)
                
                # Human-readable column headers
                l3_col_rename = {
                    's_no': '#', 'category': 'Category', 'qualification': 'Criterion',
                    'evidence_summary': 'Evidence Summary', 'reasoning': 'AI Reasoning',
                    'pass_fail': 'Pass/Fail', 'request_info': 'Need More Info?',
                    'confidence': 'Confidence'
                }
                df_crit = df_crit.rename(columns={k: v for k, v in l3_col_rename.items() if k in df_crit.columns})
                cols_to_show = ['Category', 'Criterion', 'Evidence Summary', 'AI Reasoning', 'Pass/Fail', 'Need More Info?', 'Confidence']
                existing_cols = [c for c in cols_to_show if c in df_crit.columns]
                st.dataframe(df_crit[existing_cols], use_container_width=True)
                
                if st.button("⬅️ Go Back to Step 2"):
                    st.session_state.eval_stage = 2
                    st.rerun()
                
                st.write(f"**Final Evaluated Decision:** {l3_data.get('final_decision')}")
                
                l3_count = get_precedent_count(3) + len(st.session_state.staged_precedents.get("level_3_precedents", []))
                with st.expander("✏️ Disagree with a decision? Click here to adjust", expanded=False):
                    with st.form("l3_correction_form", clear_on_submit=True):
                        st.markdown("#### Override a Pass/Fail Decision")
                        
                        if l3_count >= 45:
                            st.error(f"🚨 Correction limit reached ({l3_count}/45). Your corrections will be auto-merged. Check the sidebar to download the rulebook.")
                        elif l3_count >= 36:
                            st.warning(f"⚠️ Approaching correction limit ({l3_count}/45). Corrections will be auto-merged when the limit is reached.")
                            
                        target_crit = st.selectbox(
                            "Criterion or Decision to Adjust", 
                            [c.get('qualification', c.get('Criterion', '')) for c in criteria] + ["Final Decision"],
                            help="Select the specific criterion where the AI was too strict or too lenient."
                        )
                        correction_crit = st.selectbox(
                            "New Result", 
                            ["Pass", "Fail", "Request Additional Information"],
                            help="Select what the correct decision should be."
                        )
                        reasoning_crit = st.text_area(
                            "Why is this the correct decision?",
                            help="Explain your reasoning. The AI will remember this standard and apply it consistently in future reviews."
                        )
                        
                        is_disabled = True if l3_count >= 45 else False
                        if st.form_submit_button("Apply Correction & Re-Evaluate", disabled=is_disabled):
                            st.session_state.staged_precedents["level_3_precedents"].append({
                                "target": target_crit,
                                "correction": correction_crit,
                                "human_reasoning": reasoning_crit
                            })
                            # Re-run with FULL combined financial context (narrative PNW override preserved)
                            excel_pnw_rerun = st.session_state.l1_data.get("cross_reference_result", "None")
                            narrative_pnw_rerun = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                            combined_financials_rerun = f"Excel Cross-Reference Revenue/PNW: {excel_pnw_rerun}\nNarrative Declared PNW: {narrative_pnw_rerun}"
                            with st.spinner("Re-evaluating decisions..."):
                                st.session_state.l3_data = run_level_3_thresholds(
                                    st.session_state.l2_data.get("classifications", []), 
                                    st.session_state.l1_data.get("extracted_facts", []), 
                                    combined_financials_rerun,
                                    staged_precedents=st.session_state.staged_precedents.get("level_3_precedents", [])
                                )
                            st.rerun()
                            
                st.caption("*Your corrections are saved only after you approve the final evaluation below.*")

                if st.button("Approve Final Evaluation & Commit Corrections ➔", type="primary"):
                    from src.memory_manager import commit_staged_precedents
                    commit_staged_precedents(st.session_state.staged_precedents)
                    st.session_state.eval_stage = 4
                    st.rerun()
                    
            # STATE 4: Final Generation
            elif st.session_state.eval_stage == 4:
                st.success(f"✅ Full Process-Supervised Evaluation completed for **{file_name}**")
                
                result_md = generate_final_md_report(st.session_state.l1_data, st.session_state.l3_data)
                
                # Display Results
                st.markdown("---")
                st.markdown(result_md, unsafe_allow_html=True)
                
                colA, colB = st.columns(2)
                with colA:
                    if st.button("⬅️ Go Back to Level 3"):
                        st.session_state.eval_stage = 3
                        st.rerun()
                with colB:
                    if st.button("🔄 Reset / Start Over"):
                        for key in ['eval_stage', 'pdf_text', 'l1_data', 'l2_data', 'l3_data', 'staged_precedents', 'consolidated_rules_json', 'show_consolidation_success']:
                            if key in st.session_state:
                                del st.session_state[key]
                        st.rerun()

                # --- Handle Excel Export (Old Logic adapted for new State) ---
                from io import BytesIO
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                import re

                def parse_md_tables(md_text):
                    tables = []
                    lines = md_text.split("\n")
                    i = 0
                    while i < len(lines):
                        line = lines[i].strip()
                        if "|" in line and re.search(r"[a-zA-Z]", line):
                            header = [c.strip() for c in line.strip("|").split("|")]
                            if i + 1 < len(lines) and re.match(r"[\|\s\-:]+$", lines[i+1].strip()):
                                rows = []
                                j = i + 2
                                while j < len(lines):
                                    row_line = lines[j].strip()
                                    if "|" not in row_line or not row_line:
                                        break
                                    row = [c.strip() for c in row_line.strip("|").split("|")]
                                    rows.append(row)
                                    j += 1
                                if rows:
                                    tables.append((header, rows))
                                i = j
                                continue
                        i += 1
                    return tables

                def build_excel(md_text, report_title):
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)
                    tables = parse_md_tables(md_text)
                    sheet_names = ["Part 1 – Evaluation", "Part 2 – Explainable AI"]
                    header_fill = PatternFill("solid", fgColor="1F4E79")
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    border_side = Side(style="thin", color="000000")
                    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
                    even_fill = PatternFill("solid", fgColor="EBF3FB")

                    for t_idx, (headers, rows) in enumerate(tables):
                        sheet_name = sheet_names[t_idx] if t_idx < len(sheet_names) else f"Table {t_idx+1}"
                        ws = wb.create_sheet(title=sheet_name)
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                        title_cell = ws.cell(row=1, column=1, value=report_title)
                        title_cell.font = Font(bold=True, size=13, color="1F4E79")
                        title_cell.alignment = Alignment(horizontal="center", vertical="center")
                        ws.row_dimensions[1].height = 22
                        for col_i, h in enumerate(headers, start=1):
                            cell = ws.cell(row=2, column=col_i, value=h)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.border = cell_border
                        ws.row_dimensions[2].height = 20
                        for row_i, row in enumerate(rows, start=3):
                            fill = even_fill if row_i % 2 == 0 else PatternFill()
                            for col_i, val in enumerate(row, start=1):
                                cell = ws.cell(row=row_i, column=col_i, value=val)
                                cell.border = cell_border
                                cell.alignment = Alignment(wrap_text=True, vertical="top")
                                if fill.fill_type:
                                    cell.fill = fill
                        for col_i, h in enumerate(headers, start=1):
                            col_letter = openpyxl.utils.get_column_letter(col_i)
                            max_len = max(len(h), *[len(str(row[col_i - 1])) if col_i - 1 < len(row) else 0 for row in rows])
                            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

                    comments_match = re.search(r'### 📝 CERTIFIER COMMENTS & FINAL SUMMARY\s*(.*?)(?=\Z)', md_text, re.DOTALL | re.IGNORECASE)
                    if comments_match:
                        comments_text = comments_match.group(1).strip()
                        if comments_text:
                            ws_comments = wb.create_sheet(title="Final Summary")
                            ws_comments.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
                            title_cell = ws_comments.cell(row=1, column=1, value="Certifier Comments & Final Summary")
                            title_cell.font = Font(bold=True, size=13, color="1F4E79")
                            title_cell.alignment = Alignment(horizontal="center", vertical="center")
                            ws_comments.row_dimensions[1].height = 22
                            content_cell = ws_comments.cell(row=3, column=1, value=comments_text)
                            content_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                            ws_comments.column_dimensions['A'].width = 100
                            approx_lines = (len(comments_text) // 100) + comments_text.count("\n") + 2
                            ws_comments.row_dimensions[3].height = max(15, approx_lines * 15)

                    buf = BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    return buf

                excel_buf = build_excel(result_md, f"CUCP Evaluation — {file_name}")
                dl_col1, dl_col2 = st.columns([1, 4])
                with dl_col1:
                    st.download_button(
                        label="📊 Download (.xlsx)",
                        data=excel_buf.getvalue(),
                        file_name=f"{base_name}_evaluation.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_xlsx_0",
                    )

    elif app_option == "LLM Evaluation":
        with col22:
            st.subheader("Upload the Policy Document")
        with col24:
            knowledge_base = st.file_uploader(
                "Upload Knowledge Base",
                type=["pdf", "txt"],
                key="knowledge_base_upload",
            )
        text_based("LLM Evaluation", knowledge_base)

    elif app_option == "Foundation Model":
        foundation_model_chat_ui()
    elif app_option == "Highway Incident Summarizer Bot":
        text_based("Highway Incident Summarizer Bot", None)

    elif app_option == "LLM as a Judge":
        with col22:
            st.subheader("Upload the Policy Document")
        with col24:
            policy_file = st.file_uploader(
                "Upload Knowledge Base",
                type=["pdf", "txt"],
                key="policy_file_upload",
            )
        text_based("LLM as a Judge", policy_file)

    elif app_option == "Personal Narrative Insights":
        with col22:
            st.subheader("Upload the Personal Narrative")
        with col24:
            knowledge_base = st.file_uploader(
                "Upload Personal Narrative",
                type=["pdf", "txt"],
                key="personal_narrative_upload",
            )
        text_based("Personal Narrative Insights", None)

    elif app_option == "Highway Incident Summarizer":
        highway_incident_ui(app_option)

    else:
        with col22:
            st.subheader("Scope of Data Exchange")
            st.write("#")
        with col24:
            select_all = st.checkbox("Select All")

        base_fields_list = [
            "Actual release date",
            "Name of the youth",
            "Race/Ethnicity",
            "Medi-Cal ID Number",
            "Residential Address",
            "Telephone",
            "Medi-Cal health plan assigned",
            "Health Screenings",
            "Health Assessments",
            "Chronic Conditions",
            "Prescribed Medications",
        ]
